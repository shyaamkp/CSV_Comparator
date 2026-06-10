"""
CSV/XLSX (Single Sheet) File Comparison Tool
A standalone PyQt5 application to compare two CSV / XLS / XLSX files.

Run:
    python csv_comparator.py

Build a Windows .exe:
    pip install pyinstaller
    pyinstaller --onefile --windowed --name "CSV_Comparator" csv_comparator.py
"""

import os
import subprocess
import sys
from typing import List, Optional

import numpy as np
import pandas as pd
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QBrush, QColor, QFont, QIcon
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

PREVIEW_ROW_LIMIT = 100
COMPARISON_DISPLAY_LIMIT = 5_000
SUPPORTED_FILTER = "Spreadsheet files (*.csv *.xls *.xlsx);;CSV (*.csv);;Excel (*.xls *.xlsx);;All files (*.*)"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def read_file(path: str) -> pd.DataFrame:
    """Read a CSV / XLS / XLSX file into a DataFrame as strings (preserve formatting)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        # Try utf-8 first, fall back to latin-1 for odd encodings
        try:
            df = pd.read_csv(path, dtype=str, keep_default_na=False)
        except UnicodeDecodeError:
            df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding="latin-1")
    elif ext in (".xls", ".xlsx"):
        engine = "openpyxl" if ext == ".xlsx" else "xlrd"
        df = pd.read_excel(path, dtype=str, keep_default_na=False, engine=engine)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    df.columns = [str(c) for c in df.columns]
    return df.fillna("").astype(str)


def populate_table(table: QTableWidget, df: pd.DataFrame, limit: Optional[int] = None) -> None:
    """Render a DataFrame into a QTableWidget."""
    table.clear()
    if df is None or df.empty:
        table.setRowCount(0)
        table.setColumnCount(0)
        return

    rows = df.head(limit) if limit else df
    table.setRowCount(len(rows))
    table.setColumnCount(len(df.columns))
    table.setHorizontalHeaderLabels(list(df.columns))

    for r, (_, row) in enumerate(rows.iterrows()):
        for c, col in enumerate(df.columns):
            table.setItem(r, c, QTableWidgetItem(str(row[col])))

    table.resizeColumnsToContents()


# ---------------------------------------------------------------------------
# Multi-select column picker dialog
# ---------------------------------------------------------------------------
class ColumnPickerDialog(QDialog):
    def __init__(self, title: str, columns: List[str], selected: List[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(360, 420)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Tick the columns to include:"))

        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QAbstractItemView.NoSelection)
        for col in columns:
            item = QListWidgetItem(col)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if col in selected else Qt.Unchecked)
            self.list_widget.addItem(item)
        layout.addWidget(self.list_widget)

        btn_row = QHBoxLayout()
        select_all = QPushButton("Select All")
        clear_all = QPushButton("Clear")
        select_all.clicked.connect(lambda: self._set_all(Qt.Checked))
        clear_all.clicked.connect(lambda: self._set_all(Qt.Unchecked))
        btn_row.addWidget(select_all)
        btn_row.addWidget(clear_all)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _set_all(self, state):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(state)

    def selected_columns(self) -> List[str]:
        return [
            self.list_widget.item(i).text()
            for i in range(self.list_widget.count())
            if self.list_widget.item(i).checkState() == Qt.Checked
        ]


# ---------------------------------------------------------------------------
# Key-column mapping dialog
# ---------------------------------------------------------------------------
class KeyColumnMappingDialog(QDialog):
    """Let the user pair each key column in File 1 with any column in File 2."""

    def __init__(self, f1_cols: List[str], f2_cols: List[str], existing_mapping: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Map Key Columns")
        self.resize(520, 460)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(
            "Check each key column from File 1 and map it to the corresponding column in File 2:"
        ))

        hdr = QHBoxLayout()
        h1 = QLabel("<b>File 1 Column</b>")
        h1.setFixedWidth(210)
        hdr.addWidget(h1)
        hdr.addSpacing(24)
        hdr.addWidget(QLabel("<b>File 2 Column</b>"), 1)
        layout.addLayout(hdr)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        inner = QWidget()
        inner_layout = QVBoxLayout(inner)
        inner_layout.setSpacing(6)
        inner_layout.setContentsMargins(4, 4, 4, 4)

        self._rows: List[tuple] = []
        for col in f1_cols:
            row_w = QHBoxLayout()
            cb = QCheckBox(col)
            cb.setFixedWidth(210)
            combo = QComboBox()
            combo.addItems(f2_cols)
            mapped = existing_mapping.get(col)
            if mapped and mapped in f2_cols:
                combo.setCurrentText(mapped)
            elif col in f2_cols:
                combo.setCurrentText(col)
            cb.setChecked(col in existing_mapping)
            combo.setEnabled(col in existing_mapping)
            cb.toggled.connect(combo.setEnabled)

            arrow = QLabel("→")
            arrow.setFixedWidth(22)
            row_w.addWidget(cb)
            row_w.addWidget(arrow)
            row_w.addWidget(combo, 1)
            inner_layout.addLayout(row_w)
            self._rows.append((cb, combo))

        inner_layout.addStretch()
        scroll.setWidget(inner)
        layout.addWidget(scroll, 1)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_mapping(self) -> dict:
        return {
            cb.text(): combo.currentText()
            for cb, combo in self._rows
            if cb.isChecked()
        }


# ---------------------------------------------------------------------------
# Main window
# ---------------------------------------------------------------------------
class CSVComparator(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CSV/XLSX(Single Sheet) Comparator Tool")
        self.resize(1280, 800)

        self.file1_path: Optional[str] = None
        self.file2_path: Optional[str] = None
        self.df1: Optional[pd.DataFrame] = None
        self.df2: Optional[pd.DataFrame] = None
        self.key_columns: List[str] = []
        self.key_mapping: dict = {}
        self.ignore_columns: List[str] = []
        self.sort_column: Optional[str] = None
        self.comparison_df: Optional[pd.DataFrame] = None
        self.summary_df: Optional[pd.DataFrame] = None

        self._build_ui()

    # -- UI construction ----------------------------------------------------
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # Title
        title = QLabel("CSV/XLSX(Single Sheet) File Comparison Tool")
        title_font = QFont()
        title_font.setPointSize(13)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignCenter)
        root.addWidget(title)

        # File row
        file_row = QHBoxLayout()
        self.file1_label = QLabel("CSV/XLSX(Single Sheet) File 1: Not selected")
        self.file1_btn = QPushButton("Select CSV/XLSX(Single Sheet) 1")
        self.file1_btn.clicked.connect(lambda: self._select_file(1))
        self.file2_label = QLabel("CSV/XLSX(Single Sheet) File 2: Not selected")
        self.file2_btn = QPushButton("Select CSV/XLSX(Single Sheet) 2")
        self.file2_btn.clicked.connect(lambda: self._select_file(2))
        file_row.addWidget(self.file1_label, 1)
        file_row.addWidget(self.file1_btn, 1)
        file_row.addWidget(self.file2_label, 1)
        file_row.addWidget(self.file2_btn, 1)
        root.addLayout(file_row)

        # Comparison mode row
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("Comparison Mode"))
        self.mode_position = QRadioButton("By Position")
        self.mode_key = QRadioButton("By Key Columns")
        self.mode_position.setChecked(True)
        group = QButtonGroup(self)
        group.addButton(self.mode_position)
        group.addButton(self.mode_key)
        self.mode_position.toggled.connect(self._on_mode_change)
        mode_row.addWidget(self.mode_position)
        mode_row.addWidget(self.mode_key)
        self.key_btn = QPushButton("Select Key Columns")
        self.key_btn.clicked.connect(self._select_key_columns)
        self.key_btn.setEnabled(False)
        self.key_label = QLabel("Key Columns: None")
        mode_row.addWidget(self.key_btn)
        mode_row.addWidget(self.key_label, 1)
        root.addLayout(mode_row)

        # Ignore columns row
        ignore_row = QHBoxLayout()
        ignore_row.addWidget(QLabel("Ignore Columns:"))
        self.ignore_btn = QPushButton("Select Columns to Ignore")
        self.ignore_btn.clicked.connect(self._select_ignore_columns)
        ignore_row.addWidget(self.ignore_btn)
        self.ignore_label = QLabel("None")
        ignore_row.addWidget(self.ignore_label, 1)
        root.addLayout(ignore_row)

        # Sort row
        sort_row = QHBoxLayout()
        sort_row.addWidget(QLabel("Sort Results By:"))
        self.sort_combo = QComboBox()
        self.sort_combo.addItem("None (default order)")
        self.sort_combo.currentIndexChanged.connect(self._on_sort_change)
        sort_row.addWidget(self.sort_combo)
        sort_row.addStretch()
        root.addLayout(sort_row)

        # Tabs
        self.tabs = QTabWidget()
        root.addWidget(self.tabs, 1)
        self._build_preview_tab()
        self._build_summary_tab()
        self._build_comparison_tab()

        # Bottom action buttons
        bottom = QHBoxLayout()
        self.compare_btn = QPushButton("Compare CSV/XLSX(Single Sheet)")
        self.compare_btn.setStyleSheet(
            "background-color: #2D8CF0; color: white; padding: 10px; font-weight: bold;"
        )
        self.compare_btn.clicked.connect(self._run_comparison)

        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.setStyleSheet(
            "background-color: #19BE6B; color: white; padding: 10px; font-weight: bold;"
        )
        self.export_btn.clicked.connect(self._export_excel)

        self.reset_btn = QPushButton("⟲ Reset")
        self.reset_btn.setStyleSheet(
            "background-color: #FF9900; color: white; padding: 10px; font-weight: bold;"
        )
        self.reset_btn.clicked.connect(self._reset_all)

        bottom.addWidget(self.compare_btn, 2)
        bottom.addWidget(self.export_btn, 2)
        bottom.addWidget(self.reset_btn, 1)
        root.addLayout(bottom)

        self.statusBar().showMessage("Ready. Select two files to begin.")

    def _build_preview_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.addWidget(QLabel(f"Preview: First {PREVIEW_ROW_LIMIT} rows"))
        splitter = QSplitter(Qt.Horizontal)

        left = QWidget()
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 0, 0)
        ll.addWidget(QLabel("CSV/XLSX(Single Sheet) 1"))
        self.preview1 = QTableWidget()
        self.preview1.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        ll.addWidget(self.preview1)

        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.addWidget(QLabel("CSV/XLSX(Single Sheet) 2"))
        self.preview2 = QTableWidget()
        self.preview2.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        rl.addWidget(self.preview2)

        splitter.addWidget(left)
        splitter.addWidget(right)
        splitter.setSizes([640, 640])
        layout.addWidget(splitter)
        self.tabs.addTab(tab, "📋 Preview")

    def _build_summary_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        self.summary_table = QTableWidget()
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.summary_table)
        self.tabs.addTab(tab, "📊 Summary")

    def _build_comparison_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        legend = QHBoxLayout()
        for txt, color in [
            ("Match", "#E8F5E9"),
            ("Different value", "#FFF59D"),
            ("Only in File 1", "#FFCDD2"),
            ("Only in File 2", "#BBDEFB"),
        ]:
            sw = QLabel(f"  {txt}  ")
            sw.setStyleSheet(
                f"background-color: {color}; border: 1px solid #999; padding: 4px; margin-right: 8px;"
            )
            legend.addWidget(sw)
        legend.addStretch()
        layout.addLayout(legend)

        self.comparison_table = QTableWidget()
        self.comparison_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        layout.addWidget(self.comparison_table)
        self.tabs.addTab(tab, "🔍 Comparison Data")

    # -- Event handlers -----------------------------------------------------
    def _select_file(self, slot: int):
        path, _ = QFileDialog.getOpenFileName(
            self,
            f"Select CSV/XLSX file {slot}",
            "",
            SUPPORTED_FILTER,
        )
        if not path:
            return
        try:
            df = read_file(path)
        except Exception as e:
            QMessageBox.critical(self, "Read error", f"Failed to read file:\n{e}")
            return

        name = os.path.basename(path)
        if slot == 1:
            self.file1_path, self.df1 = path, df
            self.file1_label.setText(f"CSV/XLSX(Single Sheet) File 1: {name}")
            populate_table(self.preview1, df, PREVIEW_ROW_LIMIT)
        else:
            self.file2_path, self.df2 = path, df
            self.file2_label.setText(f"CSV/XLSX(Single Sheet) File 2: {name}")
            populate_table(self.preview2, df, PREVIEW_ROW_LIMIT)

        self._refresh_column_dependent_ui()
        self.statusBar().showMessage(f"Loaded {name} ({len(df):,} rows, {len(df.columns)} cols)")

    def _on_mode_change(self):
        self.key_btn.setEnabled(self.mode_key.isChecked())
        if not self.mode_key.isChecked():
            self.key_columns = []
            self.key_mapping = {}
            self.key_label.setText("Key Columns: None")

    def _common_columns(self) -> List[str]:
        if self.df1 is None or self.df2 is None:
            return []
        return [c for c in self.df1.columns if c in set(self.df2.columns)]

    def _refresh_column_dependent_ui(self):
        cols = self._common_columns()
        # Sort combo
        current = self.sort_combo.currentText()
        self.sort_combo.blockSignals(True)
        self.sort_combo.clear()
        self.sort_combo.addItem("None (default order)")
        for c in cols:
            self.sort_combo.addItem(c)
        idx = self.sort_combo.findText(current)
        if idx >= 0:
            self.sort_combo.setCurrentIndex(idx)
        self.sort_combo.blockSignals(False)
        # Drop stale picks
        self.ignore_columns = [c for c in self.ignore_columns if c in cols]
        f1_set = set(self.df1.columns) if self.df1 is not None else set()
        f2_set = set(self.df2.columns) if self.df2 is not None else set()
        self.key_mapping = {k: v for k, v in self.key_mapping.items() if k in f1_set and v in f2_set}
        self.key_columns = list(self.key_mapping.keys())
        self.ignore_label.setText(", ".join(self.ignore_columns) if self.ignore_columns else "None")
        self._update_key_label()

    def _select_key_columns(self):
        if self.df1 is None or self.df2 is None:
            QMessageBox.information(self, "No columns", "Load both files first.")
            return
        dlg = KeyColumnMappingDialog(
            list(self.df1.columns), list(self.df2.columns), self.key_mapping, self
        )
        if dlg.exec_() == QDialog.Accepted:
            self.key_mapping = dlg.selected_mapping()
            self.key_columns = list(self.key_mapping.keys())
            self._update_key_label()

    def _update_key_label(self):
        if not self.key_mapping:
            self.key_label.setText("Key Columns: None")
        else:
            parts = [f"{k}→{v}" if k != v else k for k, v in self.key_mapping.items()]
            self.key_label.setText("Key Columns: " + ", ".join(parts))

    def _select_ignore_columns(self):
        cols = self._common_columns()
        if not cols:
            QMessageBox.information(self, "No columns", "Load both files first.")
            return
        dlg = ColumnPickerDialog("Select Columns to Ignore", cols, self.ignore_columns, self)
        if dlg.exec_() == QDialog.Accepted:
            self.ignore_columns = dlg.selected_columns()
            self.ignore_label.setText(
                ", ".join(self.ignore_columns) if self.ignore_columns else "None"
            )

    def _on_sort_change(self, idx: int):
        self.sort_column = None if idx <= 0 else self.sort_combo.currentText()

    # -- Core comparison ----------------------------------------------------
    def _run_comparison(self):
        if self.df1 is None or self.df2 is None:
            QMessageBox.warning(self, "Missing files", "Please select both files first.")
            return
        if self.mode_key.isChecked() and not self.key_mapping:
            QMessageBox.warning(
                self, "Missing key columns", "Pick at least one key column for key-based comparison."
            )
            return

        try:
            if self.mode_position.isChecked():
                comp_df, summary = self._compare_by_position()
            else:
                comp_df, summary = self._compare_by_keys()
        except Exception as e:
            QMessageBox.critical(self, "Comparison error", str(e))
            return

        if self.sort_column and self.sort_column in comp_df.columns:
            comp_df = comp_df.sort_values(by=self.sort_column, kind="mergesort").reset_index(drop=True)

        self.comparison_df = comp_df
        self.summary_df = summary
        self._render_comparison(comp_df)
        self._render_summary(summary)
        self.tabs.setCurrentIndex(2)
        cap_note = (
            f" Showing first {COMPARISON_DISPLAY_LIMIT:,} rows — export for full results."
            if len(comp_df) > COMPARISON_DISPLAY_LIMIT else ""
        )
        self.statusBar().showMessage(
            f"Comparison complete: {len(comp_df):,} rows, "
            f"{int((comp_df['Status'] != 'MATCH').sum()):,} differences.{cap_note}"
        )

    def _columns_to_compare(self, columns: List[str]) -> List[str]:
        return [c for c in columns if c not in set(self.ignore_columns)]

    def _compare_by_position(self):
        df1, df2 = self.df1.copy(), self.df2.copy()
        common = self._columns_to_compare(self._common_columns())
        if not common:
            raise ValueError("No comparable columns between the two files.")

        n1, n2 = len(df1), len(df2)
        max_len = max(n1, n2)

        d1 = df1[common].reindex(range(max_len)).fillna("").astype(str)
        d2 = df2[common].reindex(range(max_len)).fillna("").astype(str)

        diff_mask = d1 != d2

        only1 = np.zeros(max_len, dtype=bool)
        only2 = np.zeros(max_len, dtype=bool)
        if n1 > n2:
            only1[n2:] = True
            diff_mask.iloc[n2:] = False
        elif n2 > n1:
            only2[n1:] = True
            diff_mask.iloc[n1:] = False

        has_diff = diff_mask.any(axis=1).values
        status = np.where(only1, "ONLY IN FILE 1",
                 np.where(only2, "ONLY IN FILE 2",
                 np.where(has_diff, "DIFFERENT", "MATCH")))

        differing_cols = np.full(max_len, "", dtype=object)
        diff_idx = np.where(has_diff)[0]
        if diff_idx.size:
            differing_cols[diff_idx] = (
                diff_mask.iloc[diff_idx]
                .apply(lambda row: ", ".join(c for c in common if row[c]), axis=1)
                .values
            )

        comp_df = pd.DataFrame({"Row #": np.arange(1, max_len + 1)})
        for c in common:
            comp_df[f"{c} (File1)"] = d1[c].values
            comp_df[f"{c} (File2)"] = d2[c].values
        comp_df["Status"] = status
        comp_df["Differing Columns"] = differing_cols

        summary = self._build_summary(
            int((status == "MATCH").sum()), int((status == "DIFFERENT").sum()),
            int(only1.sum()), int(only2.sum()), n1, n2, common,
        )
        return comp_df, summary

    def _compare_by_keys(self):
        df1, df2 = self.df1.copy(), self.df2.copy()

        if not self.key_mapping:
            raise ValueError("No key column mapping defined.")

        keys = list(self.key_mapping.keys())

        # Rename File 2 key columns to their File 1 counterparts so the merge aligns
        rename_f2 = {v: k for k, v in self.key_mapping.items() if k != v}
        if rename_f2:
            df2 = df2.rename(columns=rename_f2)

        # Non-key columns common to both files (by name, after key renaming)
        f2_cols = set(df2.columns)
        all_common = [c for c in df1.columns if c in f2_cols]
        compare_cols = [c for c in self._columns_to_compare(all_common) if c not in set(keys)]

        dup1 = int(df1.duplicated(subset=keys).sum())
        dup2 = int(df2.duplicated(subset=keys).sum())

        # Slice to only the columns we need, then add a within-group occurrence
        # counter so duplicate key rows are matched positionally (1st↔1st, 2nd↔2nd).
        _OCC = "__occ__"
        d1 = df1[keys + compare_cols].reset_index(drop=True)
        d2 = df2[keys + compare_cols].reset_index(drop=True)
        d1[_OCC] = d1.groupby(keys, sort=False).cumcount()
        d2[_OCC] = d2.groupby(keys, sort=False).cumcount()

        merged = pd.merge(
            d1, d2,
            on=keys + [_OCC],
            how="outer",
            suffixes=(" (File1)", " (File2)"),
            indicator=True,
        ).drop(columns=[_OCC])
        non_cat = [c for c in merged.columns if str(merged[c].dtype) != "category"]
        merged[non_cat] = merged[non_cat].fillna("")

        both = merged["_merge"] == "both"

        diff_mask = pd.DataFrame(False, index=merged.index, columns=compare_cols)
        for c in compare_cols:
            diff_mask[c] = (
                merged[f"{c} (File1)"].astype(str) != merged[f"{c} (File2)"].astype(str)
            ) & both

        has_diff = diff_mask.any(axis=1)
        status = np.where(merged["_merge"] == "left_only",  "ONLY IN FILE 1",
                 np.where(merged["_merge"] == "right_only", "ONLY IN FILE 2",
                 np.where(has_diff,                         "DIFFERENT", "MATCH")))

        differing_cols = np.full(len(merged), "", dtype=object)
        diff_idx = np.where(has_diff)[0]
        if diff_idx.size:
            differing_cols[diff_idx] = (
                diff_mask.iloc[diff_idx]
                .apply(lambda row: ", ".join(c for c in compare_cols if row[c]), axis=1)
                .values
            )

        comp_df = merged[keys].copy()
        for c in compare_cols:
            comp_df[f"{c} (File1)"] = merged[f"{c} (File1)"].values
            comp_df[f"{c} (File2)"] = merged[f"{c} (File2)"].values
        comp_df["Status"] = status
        comp_df["Differing Columns"] = differing_cols

        keys_display = [f"{k}→{v}" if k != v else k for k, v in self.key_mapping.items()]
        summary = self._build_summary(
            int((status == "MATCH").sum()), int((status == "DIFFERENT").sum()),
            int((status == "ONLY IN FILE 1").sum()), int((status == "ONLY IN FILE 2").sum()),
            len(self.df1), len(self.df2), compare_cols, keys_display,
            dup1=dup1, dup2=dup2,
        )
        return comp_df, summary

    def _build_summary(self, match, diff, only1, only2, n1, n2, compared_cols, keys=None, dup1=0, dup2=0):
        data = [
            ("File 1", os.path.basename(self.file1_path or "")),
            ("File 2", os.path.basename(self.file2_path or "")),
            ("Rows in File 1", f"{n1:,}"),
            ("Rows in File 2", f"{n2:,}"),
            ("Comparison Mode", "By Key Columns" if self.mode_key.isChecked() else "By Position"),
            ("Key Columns", ", ".join(keys) if keys else "—"),
            ("Ignored Columns", ", ".join(self.ignore_columns) if self.ignore_columns else "—"),
            ("Compared Columns", ", ".join(compared_cols) if compared_cols else "—"),
            ("Matching Rows", f"{match:,}"),
            ("Different Rows", f"{diff:,}"),
            ("Only in File 1", f"{only1:,}"),
            ("Only in File 2", f"{only2:,}"),
        ]
        if dup1 > 0 or dup2 > 0:
            data.append(("Duplicate Keys in File 1", f"{dup1:,} (first occurrence used)"))
            data.append(("Duplicate Keys in File 2", f"{dup2:,} (first occurrence used)"))
        return pd.DataFrame(data, columns=["Metric", "Value"])

    # -- Rendering ----------------------------------------------------------
    def _render_summary(self, summary: pd.DataFrame):
        populate_table(self.summary_table, summary)

    def _render_comparison(self, df: pd.DataFrame):
        display_df = df.head(COMPARISON_DISPLAY_LIMIT) if len(df) > COMPARISON_DISPLAY_LIMIT else df
        populate_table(self.comparison_table, display_df)
        if display_df.empty:
            return

        status_col = list(display_df.columns).index("Status")
        diff_col_idx = list(display_df.columns).index("Differing Columns") if "Differing Columns" in display_df.columns else -1

        color_map = {
            "MATCH": QColor("#E8F5E9"),
            "DIFFERENT": QColor("#FFF59D"),
            "ONLY IN FILE 1": QColor("#FFCDD2"),
            "ONLY IN FILE 2": QColor("#BBDEFB"),
        }

        col_index = {name: i for i, name in enumerate(display_df.columns)}

        for r in range(self.comparison_table.rowCount()):
            status_item = self.comparison_table.item(r, status_col)
            if not status_item:
                continue
            status = status_item.text()
            color = color_map.get(status, QColor("white"))
            for c in range(self.comparison_table.columnCount()):
                item = self.comparison_table.item(r, c)
                if item:
                    item.setBackground(QBrush(color))

            # Bold-highlight differing cells
            if status == "DIFFERENT" and diff_col_idx >= 0:
                diffs_item = self.comparison_table.item(r, diff_col_idx)
                if diffs_item and diffs_item.text():
                    for col_name in [s.strip() for s in diffs_item.text().split(",")]:
                        for suffix in ("(File1)", "(File2)"):
                            target = f"{col_name} {suffix}"
                            if target in col_index:
                                cell = self.comparison_table.item(r, col_index[target])
                                if cell:
                                    cell.setBackground(QBrush(QColor("#FFB300")))
                                    f = cell.font()
                                    f.setBold(True)
                                    cell.setFont(f)

    # -- Export -------------------------------------------------------------
    def _export_excel(self):
        if self.comparison_df is None:
            QMessageBox.information(self, "Nothing to export", "Run a comparison first.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save comparison report", "comparison_report.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                if self.summary_df is not None:
                    self.summary_df.to_excel(writer, sheet_name="Summary", index=False)
                self.comparison_df.to_excel(writer, sheet_name="Comparison", index=False)
                self._apply_excel_colors(writer, "Comparison", self.comparison_df)
            reply = QMessageBox.question(
                self, "Exported",
                f"Report saved to:\n{path}\n\nOpen the file now?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if reply == QMessageBox.Yes:
                if sys.platform == "win32":
                    os.startfile(path)
                elif sys.platform == "darwin":
                    subprocess.run(["open", path])
                else:
                    subprocess.run(["xdg-open", path])
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def _apply_excel_colors(self, writer, sheet: str, df: pd.DataFrame):
        from openpyxl.styles import PatternFill, Font
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter

        ws = writer.sheets[sheet]
        cols = list(df.columns)
        status_col_letter = get_column_letter(cols.index("Status") + 1)
        last_col_letter = get_column_letter(len(cols))
        data_range = f"A2:{last_col_letter}{len(df) + 1}"

        # One CF rule per status — O(1) regardless of row count
        for status_val, color in [
            ("MATCH",          "E8F5E9"),
            ("DIFFERENT",      "FFF59D"),
            ("ONLY IN FILE 1", "FFCDD2"),
            ("ONLY IN FILE 2", "BBDEFB"),
        ]:
            ws.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=[f'${status_col_letter}2="{status_val}"'],
                    fill=PatternFill(fill_type="solid", fgColor=color),
                ),
            )

        # Amber bold highlighting — only on DIFFERENT rows (small minority)
        if "Differing Columns" not in cols:
            return
        diff_fill = PatternFill("solid", fgColor="FFB300")
        bold = Font(bold=True)
        col_index = {name: i + 1 for i, name in enumerate(cols)}

        for row_idx in df.index[df["Status"] == "DIFFERENT"]:
            r = row_idx + 2  # +1 for 1-based, +1 for header
            for col_name in [s.strip() for s in str(df.at[row_idx, "Differing Columns"]).split(",") if s.strip()]:
                for suffix in ("(File1)", "(File2)"):
                    target = f"{col_name} {suffix}"
                    if target in col_index:
                        cell = ws.cell(row=r, column=col_index[target])
                        cell.fill = diff_fill
                        cell.font = bold

    # -- Reset --------------------------------------------------------------
    def _reset_all(self):
        self.file1_path = self.file2_path = None
        self.df1 = self.df2 = None
        self.key_columns = []
        self.key_mapping = {}
        self.ignore_columns = []
        self.sort_column = None
        self.comparison_df = None
        self.summary_df = None

        self.file1_label.setText("CSV/XLSX(Single Sheet) File 1: Not selected")
        self.file2_label.setText("CSV/XLSX(Single Sheet) File 2: Not selected")
        self.mode_position.setChecked(True)
        self.key_label.setText("Key Columns: None")
        self.ignore_label.setText("None")
        self.sort_combo.blockSignals(True)
        self.sort_combo.clear()
        self.sort_combo.addItem("None (default order)")
        self.sort_combo.blockSignals(False)

        for t in (self.preview1, self.preview2, self.summary_table, self.comparison_table):
            t.clear()
            t.setRowCount(0)
            t.setColumnCount(0)
        self.tabs.setCurrentIndex(0)
        self.statusBar().showMessage("Reset. Select two files to begin.")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = CSVComparator()
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
